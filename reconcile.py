import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def reconcile_olist():
    print("Ingesting Olist datasets from sample_data/...")
    orders_path = os.path.join("sample_data", "olist_orders_dataset.csv")
    payments_path = os.path.join("sample_data", "olist_order_payments_dataset.csv")

    if not os.path.exists(orders_path) or not os.path.exists(payments_path):
        raise FileNotFoundError(
            "Ensure both 'olist_orders_dataset.csv' and 'olist_order_payments_dataset.csv' are inside the 'sample_data/' directory."
        )

    # 1. Load relevant columns
    orders_df = pd.read_csv(
        orders_path, 
        usecols=["order_id", "order_status", "order_purchase_timestamp"]
    )
    payments_df = pd.read_csv(
        payments_path, 
        usecols=["order_id", "payment_sequential", "payment_type", "payment_value"]
    )

    # 2. Process all records in sample_data directly
    orders_sample = orders_df.copy()
    sample_order_ids = set(orders_sample["order_id"])

    payments_sample = payments_df[payments_df["order_id"].isin(sample_order_ids)].copy()

    # 3. Multi-rail payment aggregation (groups vouchers, credit cards, boleto per order)
    gateway_agg = payments_sample.groupby("order_id").agg(
        total_payment_collected=("payment_value", "sum"),
        payment_methods=("payment_type", lambda x: ", ".join(sorted(x.unique()))),
        transactions_count=("payment_sequential", "count")
    ).reset_index()

    # 4. Full outer join between internal orders and external settlements
    merged = pd.merge(
        orders_sample,
        gateway_agg,
        on="order_id",
        how="outer",
        indicator=True
    )

    merged["total_payment_collected"] = merged["total_payment_collected"].fillna(0)
    merged["transactions_count"] = merged["transactions_count"].fillna(0).astype(int)
    merged["payment_methods"] = merged["payment_methods"].fillna("None")

    # 5. Classify reconciliation statuses
    def classify_status(row):
        if row["_merge"] == "both":
            if row["order_status"] == "canceled" and row["total_payment_collected"] > 0:
                return "CANCELED_BUT_CHARGED"
            elif row["order_status"] == "unavailable":
                return "UNAVAILABLE_PENDING_REFUND"
            return "SETTLED_MATCH"
        elif row["_merge"] == "left_only":
            return "UNCOLLECTED_ORDER"
        elif row["_merge"] == "right_only":
            return "ORPHANED_PAYMENT"

    merged["reconciliation_status"] = merged.apply(classify_status, axis=1)

    # Clean audit dataframe
    audit_columns = [
        "order_id", "order_status", "order_purchase_timestamp",
        "total_payment_collected", "payment_methods", "transactions_count", 
        "reconciliation_status"
    ]
    audit_df = merged[audit_columns].copy()

    # 6. Calculate Executive Summary KPIs
    settled_mask = merged["reconciliation_status"] == "SETTLED_MATCH"
    canceled_charged_mask = merged["reconciliation_status"] == "CANCELED_BUT_CHARGED"
    uncollected_mask = merged["reconciliation_status"] == "UNCOLLECTED_ORDER"
    unavailable_mask = merged["reconciliation_status"] == "UNAVAILABLE_PENDING_REFUND"

    summary_data = {
        "Metric": [
            "Total Internal Orders Processed",
            "Fully Settled Orders (Clean)",
            "Canceled Orders Still Charged (Dispute Risk)",
            "Unavailable Orders Charged (Fulfillment Gap)",
            "Uncollected Orders (Zero Gateway Record)",
            "Total Revenue Reconciled (BRL)",
            "At-Risk Capital Exposure (BRL)"
        ],
        "Value": [
            len(orders_sample),
            int(settled_mask.sum()),
            int(canceled_charged_mask.sum()),
            int(unavailable_mask.sum()),
            int(uncollected_mask.sum()),
            round(float(merged.loc[settled_mask, "total_payment_collected"].sum()), 2),
            round(float(merged.loc[canceled_charged_mask | unavailable_mask, "total_payment_collected"].sum()), 2)
        ]
    }
    summary_df = pd.DataFrame(summary_data)

    # 7. Write to Excel
    os.makedirs("output", exist_ok=True)
    report_file = os.path.join("output", "Olist_RealWorld_Reconciliation.xlsx")

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Executive Summary", index=False)
        audit_df.to_excel(writer, sheet_name="Audit Trail", index=False)

    # 8. Apply conditional formatting and styling
    format_excel_report(report_file)
    print(f"Reconciliation completed successfully! Report generated at: {report_file}")

def format_excel_report(file_path):
    wb = load_workbook(file_path)

    # Style presets
    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Format Audit Trail
    ws_audit = wb["Audit Trail"]
    for cell in ws_audit[1]:
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_col_idx = [col[0].value for col in ws_audit.iter_cols()].index("reconciliation_status") + 1

    for row in range(2, ws_audit.max_row + 1):
        status = ws_audit.cell(row=row, column=status_col_idx).value
        if status in ["UNCOLLECTED_ORDER", "ORPHANED_PAYMENT"]:
            for col in range(1, ws_audit.max_column + 1):
                ws_audit.cell(row=row, column=col).fill = red_fill
        elif status in ["CANCELED_BUT_CHARGED", "UNAVAILABLE_PENDING_REFUND"]:
            for col in range(1, ws_audit.max_column + 1):
                ws_audit.cell(row=row, column=col).fill = yellow_fill

    # Format Executive Summary
    ws_summary = wb["Executive Summary"]
    for cell in ws_summary[1]:
        cell.fill = navy_fill
        cell.font = header_font

    # Auto-adjust column widths
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    wb.save(file_path)

if __name__ == "__main__":
    reconcile_olist()